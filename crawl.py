import os
import requests
import urllib.request
import pymysql
from bs4 import BeautifulSoup
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import policy
from email.utils import formataddr
from dotenv import load_dotenv


mainUrl = "https://www.saramin.co.kr"

load_dotenv(verbose=True)
SMTP_ID = os.getenv('SMTP_ID')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD')



def mariadbConnect():
    db = pymysql.connect(
        host="127.0.0.1",
        user="root",
        passwd="crawl",
        database="test",
        port=3307,
        charset="utf8",
    )
    cursor = db.cursor()
    return db, cursor


def createWorkBook():
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("생성시트")


def crawling(maxPage, cursor, db):
    for pagenumber in range(0, maxPage):
        if pagenumber == 0:
            url = (
                mainUrl
                + f"/zf_user/jobs/list/domestic?page=1&loc_mcd=101000&cat_mcls=21&isAjaxRequest=0&page_count=50&sort=RL&type=domestic&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=1#searchTitle"
            )
        url = (
            mainUrl
            + f"/zf_user/jobs/list/domestic?page={pagenumber + 1}&loc_mcd=101000&cat_mcls=21&isAjaxRequest=0&page_count=50&sort=RL&type=domestic&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=1#searchTitle"
        )
        response = requests.get(url)
        if response.status_code == 200:
            html = response.text
            soup = BeautifulSoup(html, "html.parser")
            homepage = soup.find_all("a", {"class": "str_tit"})
            indexing = 0
            for idxx, val in enumerate(homepage):
                if idxx % 2 == 0:
                    continue
                openUrl = urllib.request.urlopen(mainUrl + homepage[idxx]["href"])
                newUrl = openUrl.geturl()
                newResponse = requests.get(newUrl)
                newHtml = newResponse.text
                newsoup = BeautifulSoup(newHtml, "html.parser")
                splitList = newsoup.find_all(attrs={"name": "description"})[0][
                    "content"
                ].split(",")

                index = 0
                indexing += 1
                for idx, val in enumerate(splitList):
                    if idx == 0:
                        checksql = "SELECT * FROM company WHERE name = %s"
                        res = cursor.execute(checksql, val)
                        name = val
                        cursor.fetchall()

                    if res == 1:
                        break
                    if val.find(":") == -1:
                        continue
                    else:
                        if "홈페이지" in val:
                            tt = val.split(":", 1)
                            cursor.execute(
                                "INSERT INTO company(name, homepage) VALUES (%s, %s)",
                                (name, tt[1]),
                            )
                            cursor.fetchall()
                            db.commit()
                            print(f"수집했습니다. {name} {tt[1]}")
                        index += 1

    else:
        print(response.status_code)


def is_comapny(res):
    for i in res:
        for j in i:
            if j == "company":
                return True
    return False


def testmail(to):

    try:
        smtp = smtplib.SMTP("smtp.gmail.com", 587)
        smtp.starttls()
        smtp.login(SMTP_ID, SMTP_PASSWORD)
        msg = MIMEMultipart(policy=policy.default)
        msg["Subject"] = "사업 제안서 보냅니다."
        msg["From"] = formataddr(("찾아오는 편안함, YOGO", "contact@yogo.com"))
        msg["To"] = to
        html = """
       <body style="margin:0; padding:0; width:100%;">
      <div style="max-width:1000px; margin:0 auto;">
      <img src="https://postfiles.pstatic.net/MjAyMjAxMTBfNzIg/MDAxNjQxNzgyNTYzOTQy.-7_HTtDBTmS68BrXr0mQGYUJVZKPeWkhoFmwbksKPP4g.28ebS9YGmzuwpg3Hl7KSgK2tD5HXKxiC61JUw2wY0W4g.PNG.paristak1/KakaoTalk_20220110_113244689.png?type=w966"/>
      <div style="margin:100px auto; width:100%; text-align:center;">
      <a href="https://bit.ly/31FLxTK" style="text-align:center; text-decoration:none; font-size:40px; max-width:600px; min-width:600px; width:600px; margin:0 auto; padding:10px 20px; height:150px; background:black; color:white; border-radius:10px;">CONTACT US</a>
      </div>
      </div>
      </body>
        """
        msg.attach(MIMEText(html, "html"))
        smtp.send_message(msg)
        smtp.quit()
    except:
        f = open("log.txt", "a")
        f.write(to)


def sendMail(cursor):

    cursor.execute(
        "SELECT id, mail, isSend FROM company WHERE mail != 'None' AND isSend = 0;"
    )
    row = cursor.fetchall()
    for i in row:
        print(i)
        email = i[1]
        testmail(email)
        query = "UPDATE company SET isSend = 1 WHERE id = %s"
        cursor.execute(query, i[0])
        cursor.fetchone()


def start():
    try:
        db, cursor = mariadbConnect()
        cursor.execute("show tables")
        result = cursor.fetchall()
    except:
        return "데이터베이스를 실행시켜주세요"

    if is_comapny(result) == False:
        cursor.execute(
            "CREATE TABLE company (id INT PRIMARY KEY AUTO_INCREMENT, name VARCHAR(50) NOT NULL, homepage VARCHAR(100), mail VARCHAR(50), isSend TINYINT)"
        )
        print("테이블이 생성됐습니다.")

    action = input(
        """
    숫자 1 : 크롤링 실행
    숫자 2 : 데이터베이스 조회
    숫자 3 : 데이터베이스 엑셀로 저장
    숫자 4 : 메일 보내기 실행
    입력해주세요 : """
    )

    if action == "1":
        maxPage = input("크롤링할 페이지 수를 입력해주세요 : ")
        crawling(int(maxPage), cursor, db)
        return "finish"
    if action == "2":
        cursor.execute("SELECT * from company")
        result = cursor.fetchall()
        for row in result:
            print(row)
    if action == "3":
        cursor.execute("SELECT * from company")
        result = cursor.fetchall()
        write_wb = Workbook()
        write_st = write_wb.active

        for idx, data in enumerate(result):
            write_st.cell(row=1 + idx, column=1).value = data[1]
            write_st.cell(row=1 + idx, column=2).value = data[2]
            write_wb.save("test.xlsx")

    if action == "4":
        sendMail(cursor)
        db.commit()
        print("끝")


start()
